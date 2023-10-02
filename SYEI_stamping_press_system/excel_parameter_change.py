import openpyxl as xl
import main_program as mprog
import parameter as par

class ExcelOp(object):
    def __init__(self, file_name, sheet_name):
        self.wb = xl.load_workbook('%s.xlsx' % file_name)
        self.ws = self.wb[str(sheet_name)]

    def get_col_cell(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(2, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            if cell_value is None:
                break
            column_data.append(cell_value)
        return column_data

    def get_sheet_par(self, part_name, i):
        parameter_name = self.get_col_cell(1)
        parameter_value = self.get_col_cell(i + 2)
        if part_name != '沖床機架零件清單' and part_name != '標準資料':
            for n in range(0, len(parameter_name)):
                # 確認生成變數
                # print(parameter_name[n], parameter_value[n])
                mprog.param_change(part_name, parameter_name[n], str(parameter_value[n]))
        else:
            pass
        return parameter_name, parameter_value

    def get_assmebly_sheet_par(self, i):
        assmebly_par_list = {}
        part_name = self.get_col_cell(1)
        part_value = self.get_col_cell(i + 2)
        for x in range(len(part_name)):
            assmebly_par_list[part_name[x]] = part_value[x]
        return assmebly_par_list

    def get_assmebly_quantity(self, i):
        part_name = self.get_col_cell(1)
        part_quantity = self.get_col_cell(i + 2)
        return part_name, part_quantity

    def get_standard_parts(self, i):
        type_name = self.get_col_cell(1)
        travel_value = self.get_col_cell(2)
        close_working_height_value = self.get_col_cell(3)
        specifications_travel_min_value = self.get_col_cell(4)
        specifications_travel_max_value = self.get_col_cell(5)
        specifications_close_working_height_min_value = self.get_col_cell(6)
        specifications_close_working_height_max_value = self.get_col_cell(7)
        return type_name, travel_value, close_working_height_value, specifications_travel_min_value, specifications_travel_max_value, specifications_close_working_height_min_value, specifications_close_working_height_max_value

    def get_single_data_sheet_par(self, data_name, part_name, i):
        parameter_name = self.get_col_cell(1)
        parameter_value = self.get_col_cell(i + 2)
        position = parameter_name.index(data_name)
        if part_name != '沖床機架零件清單' and part_name != '標準資料':
            mprog.param_change(part_name, parameter_name[position], str(parameter_value[i]))
        else:
            pass
        return parameter_value

    def get_cutout_limit(self, i):
        cutout_dir = self.get_col_cell(1)
        cutout_value = self.get_col_cell(i + 2)
        for n in range(0, len(cutout_dir)):
            par.cutout_all_limit[cutout_dir[n]] = cutout_value[n]
