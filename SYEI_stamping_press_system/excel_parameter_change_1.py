import openpyxl as xl
import main_program as mprog


class ExcelOp(object):
    def __init__(self, sheet_name):
        self.wb = xl.load_workbook('尺寸整理表_1.xlsx')
        self.ws= self.wb[str(sheet_name)]

    def get_col_cell(self, column):
        rows = self.ws.max_row
        column_data = []

        for i in range(2, rows+1):
            cell_value = self.ws.cell(row=i, column=column).value
            if cell_value is None:
                break
            column_data.append(cell_value)
        return column_data

    def part_parameter(self, part_name, i):
        parameter_name = self.get_col_cell(1)
        parameter_value = self.get_col_cell(i+2)
        for n in range(0, len(parameter_name)):
            mprog.param_change(part_name, parameter_name[n], str(parameter_value[n]))

    
    def BALL_SCREW_parameter(self, BALL_SCREW_name, BALL_SCREW_number):
        parameter_name = self.get_col_cell(1)
        parameter_value = self.get_col_cell(BALL_SCREW_number+2)
        for n in range(0, len(parameter_name)):
            mprog.param_change(BALL_SCREW_name, parameter_name[n], str(parameter_value[n]))
