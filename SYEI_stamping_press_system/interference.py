import openpyxl as exl

# a = [10, 12, 30, 30, 840, 870, 80, 160, 150, 800, 5, 1390, 1207, 1218]
# b = [180, 270, 90, 270, 270, 0, 90, 90, 90, 90, 0, 180, 0, 180]
# c = 50
border = []


def interference(n, lv, a, b, c):
    # n是噸數，0~8、lv是加大型，0~2
    wb = exl.load_workbook('平板.xlsx')
    sheet1 = wb['平板']
    # 抓變數A的row
    for row_number_a, row in enumerate(sheet1.iter_rows(values_only=True), start=1):
        for column_number, cell_value in enumerate(row, start=1):
            if isinstance(cell_value, (int, str)) and cell_value == 'A':
                a_row = row_number_a
                # print(f"值 'A' 的位置在：列{column_number}，行{row_number_a}")

    # 抓變數B的row
    for row_number_b, row in enumerate(sheet1.iter_rows(values_only=True), start=1):
        for column_number, cell_value in enumerate(row, start=1):
            if isinstance(cell_value, (int, str)) and cell_value == 'B':
                b_row = row_number_b
                # print(f"值 'B' 的位置在：列{column_number}，行{row_number_b}")

    # 抓變數LV的row
    sheet2 = wb['平板尺寸']
    for row_number_lv, row in enumerate(sheet2.iter_rows(values_only=True), start=1):
        for column_number, cell_value in enumerate(row, start=1):
            if isinstance(cell_value, (int, str)) and cell_value == str("LV" + f"{lv}"):
                lv_row = row_number_lv
                # print(f"值 'LV' 的位置在：列{column_number}，行{row_number_lv}")

    # 套用平板尺寸大小
    border_a = sheet1.cell(row=a_row, column=n + 2).value + sheet2.cell(row=lv_row, column=n + 2).value
    border_b = sheet1.cell(row=b_row, column=n + 2).value

    print("0度邊界", border_a, "， 90度邊界", border_b, "\n")

    # 抓0or90度的值，儲存值到indices中
    indices_0 = [index for index, value in enumerate(b) if value == 0]
    indices_90 = [index for index, value in enumerate(b) if value == 90]
    indices_180 = [index for index, value in enumerate(b) if value == 180]
    indices_270 = [index for index, value in enumerate(b) if value == 270]
    # print(indices_0)
    # print(indices_90)

    # 偏移值加減
    for i in range(0, len(a), 1):
        border.append([a[i] + c / 2, a[i] - c / 2])
        # print(border, i)

    # 90度
    if len(indices_90) >= 1:
        # 判斷邊界
        for y1 in range(0, len(indices_90), 1):
            if border[indices_90[y1]][0] >= border_b:
                print("編號", indices_90[y1] + 1, " 左方垂直偏移", a[indices_90[y1]], "超出邊界")
            elif border[indices_90[y1]][1] <= 0:
                print("編號", indices_90[y1] + 1, " 左方垂直偏移", a[indices_90[y1]], "超出邊界")
        # 判斷干涉
        for j1 in range(1, len(indices_90), 1):
            # print(border[indices_90[j1]], j1)
            for k1 in range(0, j1, 1):
                if border[indices_90[j1]][1] <= border[indices_90[k1]][0]:
                    if border[indices_90[j1]][0] >= border[indices_90[k1]][1]:
                        print("\nerror")
                        print("編號", indices_90[k1] + 1, " 左方垂直偏移", a[indices_90[k1]])
                        print("編號", indices_90[j1] + 1, " 左方垂直偏移", a[indices_90[j1]])
                    else:
                        pass
                        # print("success")
                elif border[indices_90[j1]][1] >= border[indices_90[k1]][0]:
                    if border[indices_90[j1]][0] <= border[indices_90[k1]][1]:
                        print("\nerror")
                        print("編號", indices_90[k1] + 1, " 左方垂直偏移", a[indices_90[k1]])
                        print("編號", indices_90[j1] + 1, " 左方垂直偏移", a[indices_90[j1]])
                    else:
                        pass
                        # print("success")
                # print(border[indices_90[j1]])
                # print(border[indices_90[k1]])
        print("-----------------------------")
    # 0度
    if len(indices_0) >= 1:
        # 判斷邊界
        for x1 in range(0, len(indices_0), 1):
            if border[indices_0[x1]][0] >= border_a:
                print("編號", indices_0[x1] + 1, " 下方水平偏移", a[indices_0[x1]], "超出邊界")
            elif border[indices_0[x1]][1] <= 0:
                print("編號", indices_0[x1] + 1, " 下方水平偏移", a[indices_0[x1]], "超出邊界")
        # 判斷干涉
        for j2 in range(1, len(indices_0), 1):
            # print(border[indices_0[j2]], j2)
            for k2 in range(0, j2, 1):

                if border[indices_0[j2]][1] <= border[indices_0[k2]][0]:
                    if border[indices_0[j2]][0] >= border[indices_0[k2]][1]:
                        print("\nerror")
                        print("編號", indices_0[k2] + 1, " 下方水平偏移", a[indices_0[k2]])
                        print("編號", indices_0[j2] + 1, " 下方水平偏移", a[indices_0[j2]])
                    else:
                        pass
                        # print("success")
                elif border[indices_0[j2]][1] >= border[indices_0[k2]][0]:
                    if border[indices_0[j2]][0] <= border[indices_0[k2]][1]:
                        print("\nerror")
                        print("編號", indices_0[k2] + 1, " 下方水平偏移", a[indices_0[k2]])
                        print("編號", indices_0[j2] + 1, " 下方水平偏移", a[indices_0[j2]])
                    else:
                        pass
                        # print("success")
                # print(border[indices_0[j2]])
                # print(border[indices_0[k2]])
        print("-----------------------------")

    if len(indices_180) >= 1:
        # 判斷邊界
        for y1 in range(0, len(indices_180), 1):
            if border[indices_180[y1]][0] >= border_b:
                print("編號", indices_180[y1] + 1, " 上方水平偏移", a[indices_180[y1]], "超出邊界")
            elif border[indices_180[y1]][1] <= 0:
                print("編號", indices_180[y1] + 1, " 上方水平偏移", a[indices_180[y1]], "超出邊界")
        # 判斷干涉
        for j1 in range(1, len(indices_180), 1):
            # print(border[indices_180[j1]], j1)
            for k1 in range(0, j1, 1):
                if border[indices_180[j1]][1] <= border[indices_180[k1]][0]:
                    if border[indices_180[j1]][0] >= border[indices_180[k1]][1]:
                        print("\nerror")
                        print("編號", indices_180[k1] + 1, " 上方水平偏移", a[indices_180[k1]])
                        print("編號", indices_180[j1] + 1, " 上方水平偏移", a[indices_180[j1]])
                    else:
                        pass
                        # print("success")
                elif border[indices_180[j1]][1] >= border[indices_180[k1]][0]:
                    if border[indices_180[j1]][0] <= border[indices_180[k1]][1]:
                        print("\nerror")
                        print("編號", indices_180[k1] + 1, " 上方水平偏移", a[indices_180[k1]])
                        print("編號", indices_180[j1] + 1, " 上方水平偏移", a[indices_180[j1]])
                    else:
                        pass
                        # print("success")
                # print(border[indices_180[j1]])
                # print(border[indices_180[k1]])
        print("-----------------------------")

    if len(indices_270) >= 1:
        # 判斷邊界
        for y1 in range(0, len(indices_270), 1):
            if border[indices_270[y1]][0] >= border_a:
                print("編號", indices_270[y1] + 1, " 右方垂直偏移", a[indices_270[y1]], "超出邊界")
            elif border[indices_270[y1]][1] <= 0:
                print("編號", indices_270[y1] + 1, " 右方垂直偏移", a[indices_270[y1]], "超出邊界")
        # 判斷干涉
        for j1 in range(1, len(indices_270), 1):
            # print(border[indices_270[j1]], j1)
            for k1 in range(0, j1, 1):
                if border[indices_270[j1]][1] <= border[indices_270[k1]][0]:
                    if border[indices_270[j1]][0] >= border[indices_270[k1]][1]:
                        print("\nerror")
                        print("編號", indices_270[k1] + 1, " 右方垂直偏移", a[indices_270[k1]])
                        print("編號", indices_270[j1] + 1, " 右方垂直偏移", a[indices_270[j1]])
                    else:
                        pass
                        # print("success")
                elif border[indices_270[j1]][1] >= border[indices_270[k1]][0]:
                    if border[indices_270[j1]][0] <= border[indices_270[k1]][1]:
                        print("\nerror")
                        print("編號", indices_270[k1] + 1, " 右方垂直偏移", a[indices_270[k1]])
                        print("編號", indices_270[j1] + 1, " 右方垂直偏移", a[indices_270[j1]])
                    else:
                        pass
                        # print("success")
                # print(border[indices_270[j1]])
                # print(border[indices_270[k1]])
        print("-----------------------------")

# interference(8, 2)
